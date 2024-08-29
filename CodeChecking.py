from openpyxl import workbook, load_workbook
import CharCode

book = load_workbook('TestSheet.xlsx')
Sheet = book.active
maxrow = Sheet.max_row
PublicColumn = Sheet['C2':f'C{maxrow}']
PrivateColumn = Sheet['D2':f'D{maxrow}']


#READING PUBLIC COLUMN
for cell in PublicColumn:
    for i in cell:
        for cell in PrivateColumn:
            for k in cell:
                print(i.value, '-', k.value, '-', CharCode.KeyChecker(i.value, k.value))