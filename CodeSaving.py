from openpyxl import load_workbook
import CharCode
a = True
b = False
Book = load_workbook('TestSheet.xlsx')
Sheet = Book.active
k = Sheet.max_row
while a:
    k1 = []
    k1.append(k)

    #GETTING USER NAME
    while True:
        Name = str(input('What is your name?: '))
        sure = str(input(f'Are you sure your name is {Name}? [YES/NO]: ')).upper()
        if sure =='YES':
            k1.append(Name)
            break
        elif sure == 'NO':
            continue
        else:
            continue
        
    #GENERATING USER CODE
    UserCode = CharCode.RandCode()
    k1.append(UserCode)
    k1.append(CharCode.PrivateCode(UserCode))

    #SAVING USER INFO
    for row in Sheet[f'a{k+1}': f'd{k+1}']:
        for index, cell in enumerate(row):
            cell.value = k1[index]

    #ASKING USER IF HE WANTS TO CONTINUE
    while True:
        sure = str(input('Do you want to continue? [YES/NO]: ')).upper()
        if sure =='YES':
            k +=1
            break
        elif sure == 'NO':
            a = False
            Book.save('TestSheet.xlsx')
            break
        else:
            continue
